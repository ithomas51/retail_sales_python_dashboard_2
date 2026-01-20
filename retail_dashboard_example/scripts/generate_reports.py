"""
Sales Order Reporting Script - Charts and Workbooks
Generates marketing Excel workbook and Plotly charts from analysis data.

Usage:
    python generate_reports.py -i data/output
    python generate_reports.py --input ./analysis/sheets --output ./reports

Inputs (from analyze_sales_orders.py):
- sheets/sales_analysis_summary.csv
- sheets/sales_analysis_by_branch.csv

Outputs:
- sheets/sales_analysis_marketing.xlsx: Professional Excel workbook
- charts/*.html: Interactive Plotly charts
"""

import argparse
import logging
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# Setup logging
logger = logging.getLogger(__name__)


def create_marketing_workbook(summary_df, branch_summary, output_path):
    """Create a professionally formatted Excel workbook for marketing"""
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    money_fill = PatternFill(start_color="E8F4EA", end_color="E8F4EA", fill_type="solid")
    total_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # === Sheet 1: Executive Summary ===
    ws1 = wb.active
    if ws1 is not None:
        ws1.title = "Executive Summary"
    else:
        ws1 = wb.create_sheet("Executive Summary", 0)
    
    ws1['A1'] = "SALES ORDER ANALYSIS - EXECUTIVE SUMMARY"
    ws1['A1'].font = Font(bold=True, size=18, color="2E86AB")
    ws1.merge_cells('A1:F1')
    
    ws1['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
    ws1['A2'].font = Font(italic=True, size=10)
    
    total_row = summary_df[summary_df['Year'] == 'TOTAL'].iloc[0]
    
    metrics = [
        ("Total Line Items", f"{int(total_row['Total Line Items']):,}"),
        ("Unique Orders", f"{int(total_row['Unique Orders']):,}"),
        ("Total Revenue (After Discount)", f"${total_row['Total Revenue']:,.2f}"),
        ("Total Allowed (After Discount)", f"${total_row['Total Allowed']:,.2f}"),
        ("Retail Items", f"{int(total_row['Retail Items']):,} ({total_row['Retail %']:.1f}%)"),
        ("Insurance Items", f"{int(total_row['Insurance Items']):,} ({total_row['Insurance %']:.1f}%)"),
        ("Retail Revenue", f"${total_row['Retail Revenue']:,.2f}"),
        ("Insurance Revenue", f"${total_row['Insurance Revenue']:,.2f}"),
    ]
    
    for i, (label, value) in enumerate(metrics, start=4):
        ws1[f'A{i}'] = label
        ws1[f'A{i}'].font = Font(bold=True)
        ws1[f'B{i}'] = value
        ws1[f'B{i}'].alignment = Alignment(horizontal='right')
    
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 25
    
    # === Sheet 2: Yearly Breakdown ===
    ws2 = wb.create_sheet("Yearly Breakdown")
    
    yearly_cols = ['Year', 'Total Line Items', 'Unique Orders', 'Retail Items', 'Insurance Items',
                   'Retail %', 'Insurance %', 'Gross Allow', 'Total Discount', 'Total Revenue', 
                   'Total Allowed']
    yearly_data = summary_df[yearly_cols]
    
    for r_idx, row in enumerate(dataframe_to_rows(yearly_data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
            elif row[0] == 'TOTAL':
                cell.font = total_font
                cell.fill = total_fill
            elif c_idx >= 8:
                cell.fill = money_fill
                if isinstance(value, (int, float)):
                    cell.number_format = '$#,##0.00'
    
    for col_idx in range(1, len(yearly_cols) + 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col_idx).column_letter].width = 18
    
    # === Sheet 3: Branch Analysis ===
    ws3 = wb.create_sheet("Branch Analysis")
    
    branch_agg = branch_summary.groupby('Branch').agg({
        'Total Items': 'sum',
        'Retail Items': 'sum',
        'Insurance Items': 'sum',
        'Unique Orders': 'sum',
        'Total Revenue': 'sum',
        'Retail Revenue': 'sum',
        'Insurance Revenue': 'sum',
        'Total Allowed': 'sum'
    }).reset_index()
    
    branch_agg['Retail %'] = (branch_agg['Retail Items'] / branch_agg['Total Items'] * 100).round(2)
    branch_agg['Insurance %'] = (branch_agg['Insurance Items'] / branch_agg['Total Items'] * 100).round(2)
    
    branch_agg = branch_agg[['Branch', 'Total Items', 'Retail Items', 'Insurance Items', 'Retail %', 
                              'Insurance %', 'Unique Orders', 'Total Revenue', 'Total Allowed']]
    branch_agg = branch_agg.sort_values('Total Revenue', ascending=False)
    
    for r_idx, row in enumerate(dataframe_to_rows(branch_agg, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
            elif c_idx >= 8:
                cell.fill = money_fill
                if isinstance(value, (int, float)):
                    cell.number_format = '$#,##0.00'
    
    for col_idx in range(1, 10):
        ws3.column_dimensions[ws3.cell(row=1, column=col_idx).column_letter].width = 18
    
    # === Sheet 4: Insurance Breakdown ===
    ws4 = wb.create_sheet("Insurance Breakdown")
    
    ins_cols = ['Year', 'Primary Billing', 'Secondary Billing', 'Tertiary Billing', 'Multi-Payor Items']
    ins_data = summary_df[ins_cols]
    
    for r_idx, row in enumerate(dataframe_to_rows(ins_data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            elif row[0] == 'TOTAL':
                cell.font = total_font
                cell.fill = total_fill
    
    for col_idx in range(1, 6):
        ws4.column_dimensions[ws4.cell(row=1, column=col_idx).column_letter].width = 20
    
    wb.save(output_path)
    logger.info(f"Saved: {output_path.name}")


def create_plotly_charts(summary_df, branch_summary, output_dir):
    """Create interactive Plotly charts for branch comparison"""
    
    # Aggregate branch data
    branch_agg = branch_summary.groupby('Branch').agg({
        'Total Items': 'sum',
        'Retail Items': 'sum',
        'Insurance Items': 'sum',
        'Unique Orders': 'sum',
        'Total Revenue': 'sum',
        'Retail Revenue': 'sum',
        'Insurance Revenue': 'sum',
        'Total Allowed': 'sum'
    }).reset_index()
    
    branch_agg['Retail %'] = (branch_agg['Retail Items'] / branch_agg['Total Items'] * 100).round(2)
    branch_agg['Insurance %'] = (branch_agg['Insurance Items'] / branch_agg['Total Items'] * 100).round(2)
    branch_agg = branch_agg.sort_values('Total Revenue', ascending=True)
    
    # === Chart 1: Branch Comparison - Retail vs Insurance Items ===
    fig1 = go.Figure()
    
    fig1.add_trace(go.Bar(
        name='Retail Items',
        y=branch_agg['Branch'],
        x=branch_agg['Retail Items'],
        orientation='h',
        marker_color='#3498db',
        text=branch_agg['Retail Items'].apply(lambda x: f'{x:,.0f}'),
        textposition='inside'
    ))
    
    fig1.add_trace(go.Bar(
        name='Insurance Items',
        y=branch_agg['Branch'],
        x=branch_agg['Insurance Items'],
        orientation='h',
        marker_color='#e74c3c',
        text=branch_agg['Insurance Items'].apply(lambda x: f'{x:,.0f}'),
        textposition='inside'
    ))
    
    fig1.update_layout(
        title='Branch Comparison: Retail vs Insurance Items',
        barmode='stack',
        xaxis_title='Number of Items',
        yaxis_title='Branch',
        height=600,
        template='plotly_white',
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1)
    )
    
    fig1.write_html(output_dir / "branch_items_comparison.html")
    logger.info("Saved: branch_items_comparison.html")
    
    # === Chart 2: Branch Charges Comparison ===
    fig2 = go.Figure()
    
    fig2.add_trace(go.Bar(
        name='Retail Revenue',
        y=branch_agg['Branch'],
        x=branch_agg['Retail Revenue'],
        orientation='h',
        marker_color='#2ecc71',
        text=branch_agg['Retail Revenue'].apply(lambda x: f'${x:,.0f}'),
        textposition='inside'
    ))
    
    fig2.add_trace(go.Bar(
        name='Insurance Revenue',
        y=branch_agg['Branch'],
        x=branch_agg['Insurance Revenue'],
        orientation='h',
        marker_color='#9b59b6',
        text=branch_agg['Insurance Revenue'].apply(lambda x: f'${x:,.0f}'),
        textposition='inside'
    ))
    
    fig2.update_layout(
        title='Branch Comparison: Retail vs Insurance Revenue (After Discount)',
        barmode='stack',
        xaxis_title='Charges ($)',
        yaxis_title='Branch',
        height=600,
        template='plotly_white',
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1)
    )
    
    fig2.write_html(output_dir / "branch_charges_comparison.html")
    logger.info("Saved: branch_charges_comparison.html")
    
    # === Chart 3: Insurance vs Retail Mix by Branch (Percentage) ===
    fig3 = px.bar(
        branch_agg.sort_values('Total Items', ascending=True),
        y='Branch',
        x=['Retail %', 'Insurance %'],
        orientation='h',
        title='Branch Comparison: Retail vs Insurance Mix (%)',
        labels={'value': 'Percentage', 'variable': 'Type'},
        color_discrete_sequence=['#3498db', '#e74c3c'],
        template='plotly_white',
        height=600
    )
    fig3.update_layout(barmode='stack', legend_title_text='')
    fig3.write_html(output_dir / "branch_mix_percentage.html")
    logger.info("Saved: branch_mix_percentage.html")
    
    # === Chart 4: Year-over-Year Trend ===
    yearly_data = summary_df[summary_df['Year'] != 'TOTAL'].copy()
    yearly_data['Year'] = yearly_data['Year'].astype(str)
    
    fig4 = make_subplots(
        rows=2, cols=2,
        subplot_titles=('Total Items by Year', 'Total Revenue by Year', 
                        'Retail vs Insurance Items', 'Retail vs Insurance Revenue'),
        specs=[[{"type": "bar"}, {"type": "bar"}],
               [{"type": "bar"}, {"type": "bar"}]]
    )
    
    fig4.add_trace(go.Bar(x=yearly_data['Year'], y=yearly_data['Total Line Items'], 
                          name='Total Items', marker_color='#2E86AB'), row=1, col=1)
    
    fig4.add_trace(go.Bar(x=yearly_data['Year'], y=yearly_data['Total Revenue'], 
                          name='Total Revenue', marker_color='#A23B72'), row=1, col=2)
    
    fig4.add_trace(go.Bar(x=yearly_data['Year'], y=yearly_data['Retail Items'], 
                          name='Retail Items', marker_color='#3498db'), row=2, col=1)
    fig4.add_trace(go.Bar(x=yearly_data['Year'], y=yearly_data['Insurance Items'], 
                          name='Insurance Items', marker_color='#e74c3c'), row=2, col=1)
    
    fig4.add_trace(go.Bar(x=yearly_data['Year'], y=yearly_data['Retail Revenue'], 
                          name='Retail Revenue', marker_color='#2ecc71'), row=2, col=2)
    fig4.add_trace(go.Bar(x=yearly_data['Year'], y=yearly_data['Insurance Revenue'], 
                          name='Insurance Revenue', marker_color='#9b59b6'), row=2, col=2)
    
    fig4.update_layout(
        title='Year-over-Year Sales Analysis',
        height=800,
        template='plotly_white',
        showlegend=True,
        legend=dict(orientation='h', yanchor='bottom', y=-0.15, xanchor='center', x=0.5)
    )
    
    fig4.write_html(output_dir / "yearly_trends.html")
    logger.info("Saved: yearly_trends.html")
    
    # === Chart 5: Top Branches by Total Revenue ===
    top_branches = branch_agg.nlargest(10, 'Total Revenue')
    
    fig5 = px.bar(
        top_branches,
        x='Branch',
        y='Total Revenue',
        color='Insurance %',
        title='Top 10 Branches by Total Revenue (Color = Insurance %)',
        labels={'Total Revenue': 'Total Revenue ($)', 'Insurance %': 'Insurance %'},
        template='plotly_white',
        color_continuous_scale='RdYlGn_r',
        text=top_branches['Total Revenue'].apply(lambda x: f'${x/1e6:.1f}M')
    )
    fig5.update_traces(textposition='outside')
    fig5.update_layout(height=500)
    
    fig5.write_html(output_dir / "top_branches.html")
    logger.info("Saved: top_branches.html")


def setup_logging(log_dir: Path):
    """Configure logging to file and console"""
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file = log_dir / f"generate_reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )
    logger.info(f"Logging to: {log_file}")


def main(input_dir: Path, output_dir: Path):
    """Main function to generate reports"""
    logger.info("=" * 60)
    logger.info("SALES ORDER REPORTING")
    logger.info(f"Run Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("=" * 60)
    
    # Define paths
    sheets_dir = output_dir / "sheets"
    charts_dir = output_dir / "charts"
    
    # Ensure output directories exist
    sheets_dir.mkdir(parents=True, exist_ok=True)
    charts_dir.mkdir(parents=True, exist_ok=True)
    
    # Load source data
    summary_file = input_dir / "sales_analysis_summary.csv"
    branch_file = input_dir / "sales_analysis_by_branch.csv"
    
    if not summary_file.exists():
        logger.error(f"{summary_file} not found. Run analyze_sales_orders.py first.")
        return
    
    if not branch_file.exists():
        logger.error(f"{branch_file} not found. Run analyze_sales_orders.py first.")
        return
    
    logger.info("Loading source data...")
    summary_df = pd.read_csv(summary_file)
    branch_summary = pd.read_csv(branch_file)
    
    logger.info(f"Summary: {len(summary_df)} rows")
    logger.info(f"Branch: {len(branch_summary)} rows")
    
    # Generate Excel Workbook
    logger.info("-" * 40)
    logger.info("GENERATING MARKETING WORKBOOK")
    logger.info("-" * 40)
    excel_file = sheets_dir / "sales_analysis_marketing.xlsx"
    create_marketing_workbook(summary_df, branch_summary, excel_file)
    
    # Generate Plotly Charts
    logger.info("-" * 40)
    logger.info("GENERATING PLOTLY CHARTS")
    logger.info("-" * 40)
    create_plotly_charts(summary_df, branch_summary, charts_dir)
    
    logger.info("=" * 60)
    logger.info("REPORTING COMPLETE")
    logger.info("=" * 60)
    logger.info(f"Sheets saved to: {sheets_dir}")
    logger.info(f"Charts saved to: {charts_dir}")


def parse_args():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(
        description='Generate marketing Excel workbook and Plotly charts from analysis data'
    )
    parser.add_argument(
        '-i', '--input',
        required=True,
        help='Path to input directory containing sales_analysis_*.csv files'
    )
    parser.add_argument(
        '-o', '--output',
        default=None,
        help='Output directory for reports (default: data/output/reports)'
    )
    parser.add_argument(
        '--log-dir',
        default=None,
        help='Directory for log files (default: logs/)'
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    
    input_dir = Path(args.input)
    
    # Default output to data/output/reports if not specified
    if args.output:
        output_dir = Path(args.output)
    else:
        output_dir = Path("data/output/reports")
    
    # Setup logging
    log_dir = Path(args.log_dir) if args.log_dir else Path("logs")
    setup_logging(log_dir)
    
    # Create output directory if needed
    output_dir.mkdir(parents=True, exist_ok=True)
    
    main(input_dir, output_dir)
