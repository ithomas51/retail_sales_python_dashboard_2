"""
Invoice Reporting Script - Charts and Workbooks
Generates marketing Excel workbook and Plotly charts from invoice analysis data.

Usage:
    python generate_reports.py -i data/output
    python generate_reports.py --input ./analysis --output ./reports

Inputs (from analyze_invoices.py):
- invoice_analysis_summary.csv
- invoice_analysis_by_branch.csv
- rental_billing_analysis.csv

Outputs:
- sheets/invoice_analysis_marketing.xlsx: Professional Excel workbook
- charts/*.html: Interactive Plotly charts
"""

import argparse
import logging
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# Setup logging
logger = logging.getLogger(__name__)


def create_marketing_workbook(summary_df, branch_summary, billing_summary, output_path):
    """Create a professionally formatted Excel workbook for marketing."""
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
    money_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    total_fill = PatternFill(start_color="43A047", end_color="43A047", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # === Sheet 1: Executive Summary ===
    ws1 = wb.active
    if ws1 is not None:
        ws1.title = "Executive Summary"
    else:
        ws1 = wb.create_sheet("Executive Summary", 0)
    
    ws1['A1'] = "INVOICE ANALYSIS - EXECUTIVE SUMMARY"
    ws1['A1'].font = Font(bold=True, size=18, color="1E88E5")
    ws1.merge_cells('A1:F1')
    
    ws1['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
    ws1['A2'].font = Font(italic=True, size=10)
    
    ws1['A3'] = "5-Year Analysis (FY2021 - FY2025)"
    ws1['A3'].font = Font(bold=True, size=12)
    
    total_row = summary_df[summary_df['Year'] == 'TOTAL'].iloc[0]
    
    metrics = [
        ("Total Line Items", f"{int(total_row['Total Line Items']):,}"),
        ("Unique Invoices", f"{int(total_row['Unique Invoices']):,}"),
        ("Unique Sales Orders", f"{int(total_row['Unique Orders']):,}"),
        ("", ""),
        ("FINANCIAL METRICS", ""),
        ("Total Payments (Collected)", f"${total_row['Total Payments']:,.2f}"),
        ("Total Outstanding Balance", f"${total_row['Total Balance']:,.2f}"),
        ("Collection Rate", f"{total_row['Collection Rate %']:.1f}%"),
        ("", ""),
        ("RETAIL vs INSURANCE", ""),
        ("Retail Items (Patient)", f"{int(total_row['Retail Items']):,} ({total_row['Retail %']:.1f}%)"),
        ("Insurance Items", f"{int(total_row['Insurance Items']):,} ({total_row['Insurance %']:.1f}%)"),
        ("Retail Payments", f"${total_row['Retail Payments']:,.2f}"),
        ("Insurance Payments", f"${total_row['Insurance Payments']:,.2f}"),
        ("", ""),
        ("RENTAL METRICS", ""),
        ("New Items (Period 1)", f"{int(total_row['New Items (Period 1)']):,}"),
        ("Recurring Items (Period 2+)", f"{int(total_row['Recurring Items (Period 2+)']):,} ({total_row['Recurring %']:.1f}%)"),
        ("Avg Billing Period", f"{total_row['Avg Billing Period']:.1f} months"),
        ("Max Billing Period", f"{int(total_row['Max Billing Period'])} months"),
    ]
    
    for i, (label, value) in enumerate(metrics, start=5):
        if label == "":
            continue
        ws1[f'A{i}'] = label
        if label.isupper():
            ws1[f'A{i}'].font = Font(bold=True, size=11, color="1E88E5")
        else:
            ws1[f'A{i}'].font = Font(bold=True)
        ws1[f'B{i}'] = value
        ws1[f'B{i}'].alignment = Alignment(horizontal='right')
    
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 25
    
    # === Sheet 2: Yearly Breakdown ===
    ws2 = wb.create_sheet("Yearly Breakdown")
    
    yearly_cols = ['Year', 'Total Line Items', 'Unique Invoices', 'Retail Items', 'Insurance Items',
                   'Retail %', 'Insurance %', 'Total Payments', 'Total Balance', 'Collection Rate %',
                   'Recurring Items (Period 2+)', 'Recurring %']
    yearly_data = summary_df[[c for c in yearly_cols if c in summary_df.columns]]
    
    for r_idx, row in enumerate(dataframe_to_rows(yearly_data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
            elif row[0] == 'TOTAL':
                cell.font = total_font
                cell.fill = total_fill
            elif c_idx >= 8 and c_idx <= 9:
                cell.fill = money_fill
                if isinstance(value, (int, float)):
                    cell.number_format = '$#,##0.00'
    
    for col_idx in range(1, len(yearly_cols) + 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col_idx).column_letter].width = 18
    
    # === Sheet 3: Branch Analysis ===
    ws3 = wb.create_sheet("Branch Analysis")
    
    if branch_summary is not None and not branch_summary.empty:
        branch_agg = branch_summary.groupby('Branch').agg({
            'Total Items': 'sum',
            'Retail Items': 'sum',
            'Insurance Items': 'sum',
            'Unique Invoices': 'sum',
            'Total Payments': 'sum',
            'Retail Payments': 'sum',
            'Insurance Payments': 'sum',
            'Total Balance': 'sum',
            'Recurring Items': 'sum'
        }).reset_index()
        
        branch_agg['Retail %'] = (branch_agg['Retail Items'] / branch_agg['Total Items'] * 100).round(2)
        branch_agg['Collection Rate %'] = np.where(
            (branch_agg['Total Payments'] + branch_agg['Total Balance'].abs()) > 0,
            (branch_agg['Total Payments'] / (branch_agg['Total Payments'] + branch_agg['Total Balance'].abs()) * 100).round(2),
            100.0
        )
        
        branch_agg = branch_agg[['Branch', 'Total Items', 'Retail Items', 'Insurance Items', 'Retail %',
                                  'Unique Invoices', 'Total Payments', 'Total Balance', 'Collection Rate %']]
        branch_agg = branch_agg.sort_values('Total Payments', ascending=False)
        
        for r_idx, row in enumerate(dataframe_to_rows(branch_agg, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws3.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', wrap_text=True)
                elif c_idx >= 7 and c_idx <= 8:
                    cell.fill = money_fill
                    if isinstance(value, (int, float)):
                        cell.number_format = '$#,##0.00'
        
        for col_idx in range(1, 10):
            ws3.column_dimensions[ws3.cell(row=1, column=col_idx).column_letter].width = 18
    
    # === Sheet 4: Rental Analysis ===
    ws4 = wb.create_sheet("Rental Analysis")
    
    if billing_summary is not None and not billing_summary.empty:
        billing_agg = billing_summary.groupby('Billing Period Bucket').agg({
            'Item Count': 'sum',
            'Total Payments': 'sum',
            'Retail Items': 'sum',
            'Insurance Items': 'sum'
        }).reset_index()
        
        total_items = billing_agg['Item Count'].sum()
        total_payments = billing_agg['Total Payments'].sum()
        billing_agg['Item %'] = (billing_agg['Item Count'] / total_items * 100).round(2)
        billing_agg['Payment %'] = (billing_agg['Total Payments'] / total_payments * 100).round(2) if total_payments > 0 else 0
        
        for r_idx, row in enumerate(dataframe_to_rows(billing_agg, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws4.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
    
    wb.save(output_path)
    logger.info(f"Saved: {output_path.name}")


def create_plotly_charts(summary_df, branch_summary, billing_summary, output_dir):
    """Create interactive Plotly charts."""
    
    # Filter to non-TOTAL rows for trends
    yearly_data = summary_df[summary_df['Year'] != 'TOTAL'].copy()
    yearly_data['Year'] = yearly_data['Year'].astype(str)
    
    # === Chart 1: Branch Payments Comparison ===
    if branch_summary is not None and not branch_summary.empty:
        branch_agg = branch_summary.groupby('Branch').agg({
            'Total Payments': 'sum',
            'Retail Payments': 'sum',
            'Insurance Payments': 'sum',
            'Total Items': 'sum'
        }).reset_index()
        branch_agg = branch_agg.sort_values('Total Payments', ascending=True)
        
        fig1 = go.Figure()
        
        fig1.add_trace(go.Bar(
            name='Retail Payments',
            y=branch_agg['Branch'],
            x=branch_agg['Retail Payments'],
            orientation='h',
            marker_color='#2E7D32',
            text=branch_agg['Retail Payments'].apply(lambda x: f'${x/1000:,.0f}K' if x >= 1000 else f'${x:,.0f}'),
            textposition='inside'
        ))
        
        fig1.add_trace(go.Bar(
            name='Insurance Payments',
            y=branch_agg['Branch'],
            x=branch_agg['Insurance Payments'],
            orientation='h',
            marker_color='#1565C0',
            text=branch_agg['Insurance Payments'].apply(lambda x: f'${x/1000:,.0f}K' if x >= 1000 else f'${x:,.0f}'),
            textposition='inside'
        ))
        
        fig1.update_layout(
            title='Branch Comparison: Retail vs Insurance Payments (5-Year Total)',
            barmode='stack',
            xaxis_title='Total Payments ($)',
            yaxis_title='Branch',
            height=700,
            template='plotly_white',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1)
        )
        
        fig1.write_html(output_dir / "branch_payments_comparison.html")
        logger.info("Saved: branch_payments_comparison.html")
    
    # === Chart 2: Billing Period Distribution ===
    if billing_summary is not None and not billing_summary.empty:
        billing_agg = billing_summary.groupby('Billing Period Bucket').agg({
            'Item Count': 'sum',
            'Total Payments': 'sum'
        }).reset_index()
        
        # Sort by period order
        period_order = ['Period 1 (New)', 'Period 2-3', 'Period 4-6', 'Period 7-12', 
                        'Period 13-24', 'Period 25-36', 'Period 37+']
        billing_agg['sort_order'] = billing_agg['Billing Period Bucket'].apply(
            lambda x: period_order.index(x) if x in period_order else 99
        )
        billing_agg = billing_agg.sort_values('sort_order')
        
        fig2 = make_subplots(
            rows=1, cols=2,
            subplot_titles=('Items by Billing Period', 'Payments by Billing Period'),
            specs=[[{"type": "pie"}, {"type": "pie"}]]
        )
        
        fig2.add_trace(
            go.Pie(labels=billing_agg['Billing Period Bucket'], 
                   values=billing_agg['Item Count'],
                   hole=0.4,
                   textinfo='percent+label',
                   marker_colors=px.colors.sequential.Blues_r),
            row=1, col=1
        )
        
        fig2.add_trace(
            go.Pie(labels=billing_agg['Billing Period Bucket'], 
                   values=billing_agg['Total Payments'],
                   hole=0.4,
                   textinfo='percent+label',
                   marker_colors=px.colors.sequential.Greens_r),
            row=1, col=2
        )
        
        fig2.update_layout(
            title='Rental Billing Period Distribution (5-Year Total)',
            height=500,
            template='plotly_white'
        )
        
        fig2.write_html(output_dir / "billing_period_distribution.html")
        logger.info("Saved: billing_period_distribution.html")
    
    # === Chart 3: Collection Rate by Branch ===
    if branch_summary is not None and not branch_summary.empty:
        branch_agg = branch_summary.groupby('Branch').agg({
            'Total Payments': 'sum',
            'Total Balance': 'sum',
            'Total Items': 'sum'
        }).reset_index()
        
        branch_agg['Total Billed'] = branch_agg['Total Payments'] + branch_agg['Total Balance'].abs()
        branch_agg['Collection Rate'] = np.where(
            branch_agg['Total Billed'] > 0,
            (branch_agg['Total Payments'] / branch_agg['Total Billed'] * 100).round(1),
            100.0
        )
        branch_agg = branch_agg.sort_values('Collection Rate', ascending=True)
        
        fig3 = go.Figure()
        
        fig3.add_trace(go.Bar(
            y=branch_agg['Branch'],
            x=branch_agg['Collection Rate'],
            orientation='h',
            marker_color=branch_agg['Collection Rate'].apply(
                lambda x: '#4CAF50' if x >= 95 else ('#FFC107' if x >= 90 else '#F44336')
            ),
            text=branch_agg['Collection Rate'].apply(lambda x: f'{x:.1f}%'),
            textposition='outside'
        ))
        
        fig3.update_layout(
            title='Collection Rate by Branch (5-Year Average)',
            xaxis_title='Collection Rate (%)',
            yaxis_title='Branch',
            xaxis_range=[0, 105],
            height=700,
            template='plotly_white'
        )
        
        # Add target line
        fig3.add_vline(x=95, line_dash="dash", line_color="green", 
                       annotation_text="95% Target", annotation_position="top")
        
        fig3.write_html(output_dir / "collection_rate_by_branch.html")
        logger.info("Saved: collection_rate_by_branch.html")
    
    # === Chart 4: Yearly Trends ===
    fig4 = make_subplots(
        rows=2, cols=2,
        subplot_titles=('Total Payments by Year', 'Retail vs Insurance Payments',
                        'Collection Rate Trend', 'Recurring Revenue %'),
        specs=[[{"type": "bar"}, {"type": "bar"}],
               [{"type": "scatter"}, {"type": "scatter"}]]
    )
    
    # Total Payments
    fig4.add_trace(
        go.Bar(x=yearly_data['Year'], y=yearly_data['Total Payments'],
               name='Total Payments', marker_color='#1E88E5',
               text=yearly_data['Total Payments'].apply(lambda x: f'${x/1e6:.1f}M'),
               textposition='outside'),
        row=1, col=1
    )
    
    # Retail vs Insurance
    fig4.add_trace(
        go.Bar(x=yearly_data['Year'], y=yearly_data['Retail Payments'],
               name='Retail', marker_color='#43A047'),
        row=1, col=2
    )
    fig4.add_trace(
        go.Bar(x=yearly_data['Year'], y=yearly_data['Insurance Payments'],
               name='Insurance', marker_color='#1565C0'),
        row=1, col=2
    )
    
    # Collection Rate
    fig4.add_trace(
        go.Scatter(x=yearly_data['Year'], y=yearly_data['Collection Rate %'],
                   mode='lines+markers+text', name='Collection Rate',
                   line=dict(color='#FF9800', width=3),
                   text=yearly_data['Collection Rate %'].apply(lambda x: f'{x:.1f}%'),
                   textposition='top center'),
        row=2, col=1
    )
    
    # Recurring %
    fig4.add_trace(
        go.Scatter(x=yearly_data['Year'], y=yearly_data['Recurring %'],
                   mode='lines+markers+text', name='Recurring %',
                   line=dict(color='#9C27B0', width=3),
                   text=yearly_data['Recurring %'].apply(lambda x: f'{x:.1f}%'),
                   textposition='top center'),
        row=2, col=2
    )
    
    fig4.update_layout(
        title='Year-over-Year Invoice Analysis (5-Year)',
        height=800,
        template='plotly_white',
        showlegend=True,
        legend=dict(orientation='h', yanchor='bottom', y=-0.15, xanchor='center', x=0.5)
    )
    
    fig4.write_html(output_dir / "yearly_trends.html")
    logger.info("Saved: yearly_trends.html")
    
    # === Chart 5: Top Branches by Payments ===
    if branch_summary is not None and not branch_summary.empty:
        branch_agg = branch_summary.groupby('Branch').agg({
            'Total Payments': 'sum',
            'Retail Items': 'sum',
            'Total Items': 'sum'
        }).reset_index()
        branch_agg['Retail %'] = (branch_agg['Retail Items'] / branch_agg['Total Items'] * 100).round(1)
        top_branches = branch_agg.nlargest(10, 'Total Payments')
        
        fig5 = px.bar(
            top_branches,
            x='Branch',
            y='Total Payments',
            color='Retail %',
            title='Top 10 Branches by Total Payments (Color = Retail %)',
            labels={'Total Payments': 'Total Payments ($)', 'Retail %': 'Retail %'},
            template='plotly_white',
            color_continuous_scale='RdYlGn',
            text=top_branches['Total Payments'].apply(lambda x: f'${x/1e6:.1f}M' if x >= 1e6 else f'${x/1e3:.0f}K')
        )
        fig5.update_traces(textposition='outside')
        fig5.update_layout(height=500)
        
        fig5.write_html(output_dir / "top_branches.html")
        logger.info("Saved: top_branches.html")


def setup_logging(log_dir: Path):
    """Configure logging to file and console."""
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file = log_dir / f"generate_reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )
    logger.info(f"Logging to: {log_file}")


def main(input_dir: Path, output_dir: Path):
    """Main function to generate reports."""
    logger.info("=" * 60)
    logger.info("INVOICE REPORTING")
    logger.info(f"Run Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("=" * 60)
    
    # Define paths
    sheets_dir = output_dir / "sheets"
    charts_dir = output_dir / "charts"
    
    sheets_dir.mkdir(parents=True, exist_ok=True)
    charts_dir.mkdir(parents=True, exist_ok=True)
    
    # Load source data
    summary_file = input_dir / "invoice_analysis_summary.csv"
    branch_file = input_dir / "invoice_analysis_by_branch.csv"
    billing_file = input_dir / "rental_billing_analysis.csv"
    
    if not summary_file.exists():
        logger.error(f"{summary_file} not found. Run analyze_invoices.py first.")
        return
    
    logger.info("Loading source data...")
    summary_df = pd.read_csv(summary_file)
    
    branch_summary = None
    if branch_file.exists():
        branch_summary = pd.read_csv(branch_file)
        logger.info(f"Branch data: {len(branch_summary)} rows")
    
    billing_summary = None
    if billing_file.exists():
        billing_summary = pd.read_csv(billing_file)
        logger.info(f"Billing data: {len(billing_summary)} rows")
    
    logger.info(f"Summary: {len(summary_df)} rows")
    
    # Generate Excel Workbook
    logger.info("-" * 40)
    logger.info("GENERATING MARKETING WORKBOOK")
    logger.info("-" * 40)
    excel_file = sheets_dir / "invoice_analysis_marketing.xlsx"
    create_marketing_workbook(summary_df, branch_summary, billing_summary, excel_file)
    
    # Generate Plotly Charts
    logger.info("-" * 40)
    logger.info("GENERATING PLOTLY CHARTS")
    logger.info("-" * 40)
    create_plotly_charts(summary_df, branch_summary, billing_summary, charts_dir)
    
    logger.info("=" * 60)
    logger.info("REPORTING COMPLETE")
    logger.info("=" * 60)
    logger.info(f"Sheets saved to: {sheets_dir}")
    logger.info(f"Charts saved to: {charts_dir}")


def parse_args():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description='Generate marketing Excel workbook and Plotly charts from invoice analysis'
    )
    parser.add_argument(
        '-i', '--input',
        required=True,
        help='Path to input directory containing invoice_analysis_*.csv files'
    )
    parser.add_argument(
        '-o', '--output',
        default=None,
        help='Output directory for reports (default: data/output/reports)'
    )
    parser.add_argument(
        '--log-dir',
        default=None,
        help='Directory for log files (default: logs/)'
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    
    input_dir = Path(args.input)
    
    if args.output:
        output_dir = Path(args.output)
    else:
        output_dir = Path("data/output/reports")
    
    log_dir = Path(args.log_dir) if args.log_dir else Path("logs")
    setup_logging(log_dir)
    
    output_dir.mkdir(parents=True, exist_ok=True)
    
    main(input_dir, output_dir)
